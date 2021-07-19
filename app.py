import openpyxl as xl
from openpyxl.chart import BarChart, Reference


workbook = xl.load_workbook("transactions.xlsx")
sheet = workbook["Sheet1"]

for row in range(2, sheet.max_row +1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value*0.9

    new_cell = sheet.cell(row, 4)
    new_cell.value = corrected_price

sheet.cell(1,4).value = "New Price (- 10%)"

values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4
            )

chart = BarChart()
chart.add_data(values)

sheet.add_chart(chart, "A7")
workbook.save("output_transaction.xlsx")
    

